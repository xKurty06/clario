from pathlib import Path

import openpyxl
from fastapi import UploadFile
from fastapi.testclient import TestClient

from app.database.migrations import migrate
from app.main import app
from app.repositories.session_repository import SessionRepository
from app.services import file_service, sheet_service


def register_file(file_id: str, path: Path) -> None:
    file_service._files[file_id] = (path.name, path, path.stat().st_size)


def make_preview_workbook(path: Path) -> None:
    book = openpyxl.Workbook()
    sheet = book.active
    assert sheet is not None
    sheet.title = "Items"
    sheet.append(["Item Number", "Description", "Quantity"])
    sheet.append(["1", "Bond Paper", 2])
    book.save(path)


def preview_payload(file_id: str, file_name: str) -> dict[str, object]:
    return {
        "data_source": {
            "id": "source-1",
            "name": "Items source",
            "file_id": file_id,
            "file_name": file_name,
            "sheet_name": "Items",
            "header_row": 1,
            "first_data_row": 2,
            "selected_row_numbers": [],
            "ignored_row_numbers": [],
            "row_selection_mode": "auto_detected",
            "fields": [],
        }
    }


def test_data_source_preview_route_returns_preview(tmp_path: Path) -> None:
    path = tmp_path / "preview.xlsx"
    make_preview_workbook(path)
    register_file("file-1", path)

    response = TestClient(app).post(
        "/api/v1/files/data-source-preview",
        json=preview_payload("file-1", "preview.xlsx"),
    )

    assert response.status_code == 200
    body = response.json()
    data_row = next(row for row in body["rows"] if row["row_number"] == 2)
    assert body["data_source"]["selected_row_numbers"] == [2]
    assert body["columns"][1]["letter"] == "B"
    assert data_row["cells"]["Description"] == "Bond Paper"


def test_data_source_preview_keeps_uploaded_display_name(tmp_path: Path) -> None:
    path = tmp_path / "abc123.xlsx"
    make_preview_workbook(path)
    register_file("file-1", path)

    response = TestClient(app).post(
        "/api/v1/files/data-source-preview",
        json=preview_payload("file-1", "original-upload-name.xlsx"),
    )

    assert response.status_code == 200
    body = response.json()
    assert body["data_source"]["file_name"] == "original-upload-name.xlsx"


def test_inspect_sheets_returns_fast_basic_metadata(tmp_path: Path) -> None:
    path = tmp_path / "multi-sheet.xlsx"
    book = openpyxl.Workbook()
    first = book.active
    assert first is not None
    first.title = "Cover"
    first.append(["Document title"])
    first.append(["Item No", "Description", "Qty"])
    first.append([1, "Bond Paper", 2])
    second = book.create_sheet("Other")
    second.append(["Header A", "Header B"])
    book.save(path)

    sheets = sheet_service.inspect_sheets(path)

    assert [sheet.name for sheet in sheets] == ["Cover", "Other"]
    assert sheets[0].detected_header_row == 2
    assert sheets[0].headers == []
    assert sheets[0].sample_rows == []


def test_get_file_recovers_persisted_upload_metadata(tmp_path: Path, monkeypatch) -> None:
    upload_dir = tmp_path / "working-files"
    upload_dir.mkdir()
    path = upload_dir / "abc123.xlsx"
    make_preview_workbook(path)

    monkeypatch.setattr(file_service.get_settings(), "data_directory", tmp_path)
    file_service._files.clear()
    file_service._write_file_metadata("abc123", "original.xlsx", path, path.stat().st_size)
    file_service._files.clear()

    name, recovered_path, size = file_service.get_file("abc123")

    assert name == "original.xlsx"
    assert recovered_path == path
    assert size == path.stat().st_size


def test_get_file_recovers_existing_disk_file_without_metadata(tmp_path: Path, monkeypatch) -> None:
    upload_dir = tmp_path / "working-files"
    upload_dir.mkdir()
    path = upload_dir / "abc123.xlsx"
    make_preview_workbook(path)

    monkeypatch.setattr(file_service.get_settings(), "data_directory", tmp_path)
    file_service._files.clear()

    name, recovered_path, size = file_service.get_file("abc123")

    assert name == "abc123.xlsx"
    assert recovered_path == path
    assert size == path.stat().st_size


def test_session_draft_persists_all_uploaded_files(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(file_service.get_settings(), "data_directory", tmp_path)
    file_service._files.clear()
    migrate()

    first_path = tmp_path / "first.csv"
    second_path = tmp_path / "second.csv"
    first_path.write_text("col\n1\n", encoding="utf-8")
    second_path.write_text("col\n2\n", encoding="utf-8")

    first_upload = UploadFile(filename="first.csv", file=first_path.open("rb"))
    second_upload = UploadFile(filename="second.csv", file=second_path.open("rb"))

    first_id, first_name, _, _ = __import__("asyncio").run(file_service.save_upload(first_upload))
    second_id, second_name, _, _ = __import__("asyncio").run(file_service.save_upload(second_upload))

    created = SessionRepository().create_draft(
        "Multi File Session",
        "custom_comparison_builder",
        [first_name, second_name],
        [first_id, second_id],
    )

    state = SessionRepository().get_state(created["id"])
    assert state is not None
    assert [file["name"] for file in state["files"]] == [first_name, second_name]
    assert [file["id"] for file in state["files"]] == [first_id, second_id]


def test_session_remove_file_deletes_persisted_upload_and_metadata(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(file_service.get_settings(), "data_directory", tmp_path)
    file_service._files.clear()
    migrate()

    workbook_path = tmp_path / "remove-me.csv"
    workbook_path.write_text("col\n1\n", encoding="utf-8")
    upload = UploadFile(filename="remove-me.csv", file=workbook_path.open("rb"))
    file_id, file_name, _, _ = __import__("asyncio").run(file_service.save_upload(upload))

    created = SessionRepository().create_draft("Remove File Session", "custom_comparison_builder", [file_name], [file_id])
    session_path = SessionRepository().get_session_directory(created["id"])
    assert session_path is not None
    uploaded_copy = next((session_path / "uploads").glob(f"{file_id}.*"))
    assert uploaded_copy.exists()

    removed = SessionRepository().remove_file(created["id"], file_id)

    assert removed is True
    assert not any((session_path / "uploads").glob(f"{file_id}.*"))
    state = SessionRepository().get_state(created["id"])
    assert state is not None
    assert [file["name"] for file in state["files"]] == []
