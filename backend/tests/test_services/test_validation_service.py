from pathlib import Path

import openpyxl

from app.models.comparison_models import ComparisonDataSource, ComparisonField, ComparisonRule, ValidationRequest
from app.services import file_service
from app.services.validation_service import run_validation


def register_file(file_id: str, path: Path) -> None:
    file_service._files[file_id] = (path.name, path, path.stat().st_size)


def make_field(data_source_id: str, name: str, letter: str, field_type: str = "text", required: bool = False) -> ComparisonField:
    return ComparisonField(
        id=f"{data_source_id}-{name.lower().replace(' ', '-')}",
        data_source_id=data_source_id,
        field_name=name,
        field_type=field_type,
        column_letter=letter,
        original_header_label=name,
        custom_display_name=name,
        required=required,
    )


def test_rule_based_validation_supports_compare_formula_required_and_duplicate(tmp_path: Path) -> None:
    path = tmp_path / "builder.xlsx"
    book = openpyxl.Workbook()
    reference = book.active
    reference.title = "Reference"
    reference.append(["Item Number", "Description", "Quantity"])
    reference.append(["1", "Bond Paper", 2])
    reference.append(["1", "Bond Paper", 2])

    comparison = book.create_sheet("Comparison")
    comparison.append(["Item Number", "Description", "Quantity"])
    comparison.append(["1", "Bond Paper A4", 3])

    bidder = book.create_sheet("Bidder")
    bidder.append(["Quantity", "Unit Cost", "Total Cost", "Notes"])
    bidder.append([2, 100, 150, ""])
    book.save(path)

    register_file("reference", path)
    register_file("comparison", path)
    register_file("bidder", path)

    reference_source = ComparisonDataSource(
        id="reference-source",
        name="Reference",
        file_id="reference",
        sheet_name="Reference",
        header_row=1,
        first_data_row=2,
        selected_row_numbers=[2, 3],
        fields=[
            make_field("reference-source", "Item Number", "A"),
            make_field("reference-source", "Description", "B"),
            make_field("reference-source", "Quantity", "C", "number"),
        ],
    )
    comparison_source = ComparisonDataSource(
        id="comparison-source",
        name="Comparison",
        file_id="comparison",
        sheet_name="Comparison",
        header_row=1,
        first_data_row=2,
        selected_row_numbers=[2],
        fields=[
            make_field("comparison-source", "Item Number", "A"),
            make_field("comparison-source", "Description", "B"),
            make_field("comparison-source", "Quantity", "C", "number"),
        ],
    )
    bidder_source = ComparisonDataSource(
        id="bidder-source",
        name="Bidder",
        file_id="bidder",
        sheet_name="Bidder",
        header_row=1,
        first_data_row=2,
        selected_row_numbers=[2],
        fields=[
            make_field("bidder-source", "Quantity", "A", "number"),
            make_field("bidder-source", "Unit Cost", "B", "currency"),
            make_field("bidder-source", "Total Cost", "C", "currency"),
            make_field("bidder-source", "Notes", "D", "text", required=True),
        ],
    )

    request = ValidationRequest(
        project_name="Builder validation",
        preset="reference_bidder_abstract",
        data_sources=[reference_source, comparison_source, bidder_source],
        rules=[
          ComparisonRule(
              id="description-rule",
              rule_name="Description compare",
              rule_type="compare_values",
              left_data_source_id="reference-source",
              left_field_id="reference-source-description",
              right_data_source_id="comparison-source",
              right_field_id="comparison-source-description",
              left_match_field_ids=["reference-source-item-number"],
              right_match_field_ids=["comparison-source-item-number"],
              match_strategy="by_item_number_field",
              strictness="normalized_exact",
              severity="high",
          ),
          ComparisonRule(
              id="formula-rule",
              rule_name="Bidder total formula",
              rule_type="formula_check",
              left_data_source_id="bidder-source",
              left_field_id="bidder-source-total-cost",
              strictness="currency_tolerance",
              currency_tolerance="0.01",
              formula_settings={"operator": "multiply", "operand_field_ids": ["bidder-source-quantity", "bidder-source-unit-cost"], "result_field_id": "bidder-source-total-cost"},
              severity="high",
          ),
          ComparisonRule(
              id="required-rule",
              rule_name="Bidder notes required",
              rule_type="required_field_check",
              left_data_source_id="bidder-source",
              left_field_id="bidder-source-notes",
              severity="medium",
          ),
          ComparisonRule(
              id="duplicate-rule",
              rule_name="Reference item unique",
              rule_type="duplicate_check",
              left_data_source_id="reference-source",
              left_field_id="reference-source-item-number",
              left_match_field_ids=["reference-source-item-number"],
              severity="medium",
          ),
        ],
    )

    result = run_validation(request)

    assert result.total_selected_rows == 4
    assert {item.rule_id for item in result.discrepancies} >= {"description-rule", "formula-rule", "required-rule", "duplicate-rule"}


def test_blank_match_key_rows_do_not_create_redundant_missing_match_discrepancies(tmp_path: Path) -> None:
    path = tmp_path / "blank-row.xlsx"
    book = openpyxl.Workbook()
    reference = book.active
    reference.title = "Reference"
    reference.append(["SKU", "Description", "Quantity", "Notes"])
    reference.append(["A-1", "Bond Paper", 2])
    reference.append([None, None, None, "Manually selected non-data row"])

    comparison = book.create_sheet("Comparison")
    comparison.append(["SKU", "Description", "Quantity", "Notes"])
    comparison.append(["A-1", "Bond Paper", 2])
    book.save(path)

    register_file("blank-reference", path)
    register_file("blank-comparison", path)

    reference_source = ComparisonDataSource(
        id="blank-reference-source",
        name="Reference",
        file_id="blank-reference",
        sheet_name="Reference",
        header_row=1,
        first_data_row=2,
        row_selection_mode="manual_include",
        selected_row_numbers=[2, 3],
        fields=[
            make_field("blank-reference-source", "SKU", "A", required=True),
            make_field("blank-reference-source", "Description", "B", required=True),
            make_field("blank-reference-source", "Quantity", "C", "number", required=True),
        ],
    )
    comparison_source = ComparisonDataSource(
        id="blank-comparison-source",
        name="Comparison",
        file_id="blank-comparison",
        sheet_name="Comparison",
        header_row=1,
        first_data_row=2,
        row_selection_mode="manual_include",
        selected_row_numbers=[2],
        fields=[
            make_field("blank-comparison-source", "SKU", "A", required=True),
            make_field("blank-comparison-source", "Description", "B", required=True),
            make_field("blank-comparison-source", "Quantity", "C", "number", required=True),
        ],
    )
    request = ValidationRequest(
        project_name="Blank row validation",
        data_sources=[reference_source, comparison_source],
        rules=[
            ComparisonRule(
                id="quantity-rule",
                rule_name="Quantity comparison",
                rule_type="compare_values",
                left_data_source_id="blank-reference-source",
                left_field_id="blank-reference-source-quantity",
                right_data_source_id="blank-comparison-source",
                right_field_id="blank-comparison-source-quantity",
                left_match_field_ids=["blank-reference-source-sku"],
                right_match_field_ids=["blank-comparison-source-sku"],
                match_strategy="by_item_number_field",
                strictness="numeric_tolerance",
                numeric_tolerance="0",
                severity="medium",
            )
        ],
    )

    result = run_validation(request)

    assert any(item.rule_id == "extraction" and item.left_row_number == 3 for item in result.discrepancies)
    assert not any(
        item.rule_id == "quantity-rule" and item.left_row_number == 3 and item.actual_value == "Missing right match"
        for item in result.discrepancies
    )


def test_same_row_match_key_typo_reports_one_discrepancy_with_both_locations(tmp_path: Path) -> None:
    path = tmp_path / "match-key-typo.xlsx"
    book = openpyxl.Workbook()
    reference = book.active
    reference.title = "Reference"
    reference.append(["SKU", "Description"])
    reference.append(["INK-MRK-BLK", "Industrial permanent marker, black"])

    comparison = book.create_sheet("Comparison")
    comparison.append(["SKU", "Description"])
    comparison.append(["INK-MRK-BLCK", "Industrial permanent marker, black"])
    book.save(path)

    register_file("typo-reference", path)
    register_file("typo-comparison", path)

    reference_source = ComparisonDataSource(
        id="typo-reference-source",
        name="Reference",
        file_id="typo-reference",
        sheet_name="Reference",
        header_row=1,
        first_data_row=2,
        row_selection_mode="manual_include",
        selected_row_numbers=[2],
        fields=[
            make_field("typo-reference-source", "SKU", "A"),
            make_field("typo-reference-source", "Description", "B"),
        ],
    )
    comparison_source = ComparisonDataSource(
        id="typo-comparison-source",
        name="Comparison",
        file_id="typo-comparison",
        sheet_name="Comparison",
        header_row=1,
        first_data_row=2,
        row_selection_mode="manual_include",
        selected_row_numbers=[2],
        fields=[
            make_field("typo-comparison-source", "SKU", "A"),
            make_field("typo-comparison-source", "Description", "B"),
        ],
    )
    request = ValidationRequest(
        project_name="Typo validation",
        data_sources=[reference_source, comparison_source],
        rules=[
            ComparisonRule(
                id="description-rule",
                rule_name="Description comparison",
                rule_type="compare_values",
                left_data_source_id="typo-reference-source",
                left_field_id="typo-reference-source-description",
                right_data_source_id="typo-comparison-source",
                right_field_id="typo-comparison-source-description",
                left_match_field_ids=["typo-reference-source-sku"],
                right_match_field_ids=["typo-comparison-source-sku"],
                match_strategy="by_item_number_field",
                strictness="normalized_exact",
                severity="high",
            )
        ],
    )

    result = run_validation(request)
    description_discrepancies = [item for item in result.discrepancies if item.rule_id == "description-rule"]

    assert len(description_discrepancies) == 1
    assert description_discrepancies[0].left_row_number == 2
    assert description_discrepancies[0].right_row_number == 2
    assert description_discrepancies[0].expected_value == "INK-MRK-BLK"
    assert description_discrepancies[0].actual_value == "INK-MRK-BLCK"
