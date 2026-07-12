from pathlib import Path

import openpyxl

from app.models.comparison_models import ComparisonDataSource, ComparisonField
from app.services import file_service
from app.services.extraction_service import extract_records, preview_data_source


def register_file(file_id: str, path: Path) -> None:
    file_service._files[file_id] = (path.name, path, path.stat().st_size)


def test_preview_and_extraction_respect_selected_and_ignored_rows(tmp_path: Path) -> None:
    path = tmp_path / "rows.xlsx"
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Items"
    sheet.append(["Item Number", "Description", "Quantity"])
    sheet.append(["1", "Bond Paper", 2])
    sheet.append(["2", "Notebook", 5])
    sheet.append(["3", "Marker", 7])
    book.save(path)
    register_file("file-1", path)

    source = ComparisonDataSource(
        id="source-1",
        name="Items source",
        file_id="file-1",
        sheet_name="Items",
        header_row=1,
        first_data_row=2,
        selected_row_numbers=[2, 3, 4],
        ignored_row_numbers=[3],
        fields=[
            ComparisonField(id="field-description", data_source_id="source-1", field_name="Description", field_type="text", column_letter="B"),
            ComparisonField(id="field-quantity", data_source_id="source-1", field_name="Quantity", field_type="number", column_letter="C"),
        ],
    )

    preview = preview_data_source(source)
    records = extract_records(source)

    assert preview.columns[1].letter == "B"
    assert [record.excel_row_number for record in records] == [2, 4]
    assert records[0].field_values["field-description"].raw_value == "Bond Paper"
    assert str(records[1].field_values["field-quantity"].normalized_value) == "7"


def test_preview_excludes_rows_before_first_data_row(tmp_path: Path) -> None:
    path = tmp_path / "offset-rows.xlsx"
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Items"
    sheet.append(["Cycle: 2026-Q2"])
    sheet.append(["Distribution Hub"])
    sheet.append([])
    sheet.append(["Planning Values"])
    sheet.append(["Line", "SKU", "Description"])
    sheet.append(["ZONE A", "ZONE A", "ZONE A"])
    sheet.append([1, "CLD-GEL-500", "Reusable gel pack"])
    sheet.append([2, "CLD-LIN-032", "Thermal box liner"])
    book.save(path)
    register_file("file-offset", path)

    source = ComparisonDataSource(
        id="source-offset",
        name="Offset source",
        file_id="file-offset",
        sheet_name="Items",
        header_row=5,
        first_data_row=7,
        selected_row_numbers=[2, 4, 6, 7, 8],
        ignored_row_numbers=[],
        fields=[],
    )

    preview = preview_data_source(source)
    records = extract_records(source)

    assert preview.rows[0].row_number == 1
    assert preview.data_source.selected_row_numbers == [7, 8]
    assert preview.data_source.ignored_row_numbers == [1, 2, 3, 4, 6]
    assert [row.row_number for row in preview.rows if row.selected] == [7, 8]
    assert [row.row_number for row in preview.rows if row.ignored] == [1, 2, 3, 4, 6]
    assert [record.excel_row_number for record in records] == [7, 8]


def test_preview_leaves_repeated_section_rows_unselected_but_not_ignored(tmp_path: Path) -> None:
    path = tmp_path / "section-rows.xlsx"
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Items"
    sheet.append(["Line", "SKU", "Description", "Pack Size", "UOM"])
    sheet.append([1, "CLD-GEL-500", "Reusable gel pack", "24 pcs / carton", "carton"])
    sheet.append([2, "CLD-BOX-032", "Thermal box insert", "50 sets / bundle", "bundle"])
    sheet.append(["ZONE B - WAREHOUSE CONSUMABLES"] * 5)
    sheet.append([3, "WRP-STRETCH-18", "Stretch film roll", "6 rolls / case", "case"])
    book.save(path)
    register_file("file-section", path)

    source = ComparisonDataSource(
        id="source-section",
        name="Section source",
        file_id="file-section",
        sheet_name="Items",
        header_row=1,
        first_data_row=2,
        selected_row_numbers=[],
        ignored_row_numbers=[],
        fields=[],
    )

    preview = preview_data_source(source)
    records = extract_records(source)

    assert preview.data_source.selected_row_numbers == [2, 3, 5]
    assert preview.data_source.ignored_row_numbers == []
    assert [row.row_number for row in preview.rows if row.ignored] == []
    assert [row.row_number for row in preview.rows if row.selected] == [2, 3, 5]
    assert [record.excel_row_number for record in records] == [2, 3, 5]


def test_preview_does_not_treat_header_words_inside_data_as_header_rows(tmp_path: Path) -> None:
    path = tmp_path / "line-substring.xlsx"
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Items"
    sheet.append(["Line", "SKU", "Description", "Pack Size", "UOM"])
    sheet.append([1, "CLD-GEL-500", "Reusable gel pack", "24 pcs / carton", "carton"])
    sheet.append([2, "CLD-LIN-032", "Thermal box liner", "50 sets / bundle", "bundle"])
    sheet.append([3, "CLD-TMP-IND", "Temperature indicator strip", "100 strips / pack", "pack"])
    book.save(path)
    register_file("file-line-substring", path)

    source = ComparisonDataSource(
        id="source-line-substring",
        name="Line substring source",
        file_id="file-line-substring",
        sheet_name="Items",
        header_row=1,
        first_data_row=2,
        selected_row_numbers=[],
        ignored_row_numbers=[],
        fields=[],
    )

    preview = preview_data_source(source)

    assert preview.data_source.selected_row_numbers == [2, 3, 4]


def test_auto_detected_preview_ignores_stale_selected_rows(tmp_path: Path) -> None:
    path = tmp_path / "stale-selected.xlsx"
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Items"
    sheet.append(["Line", "SKU", "Description", "Pack Size", "UOM"])
    sheet.append(["ZONE A"] * 5)
    sheet.append([1, "CLD-GEL-500", "Reusable gel pack", "24 pcs / carton", "carton"])
    sheet.append([2, "CLD-LIN-032", "Thermal box liner", "50 sets / bundle", "bundle"])
    sheet.append([3, "CLD-TMP-IND", "Temperature indicator strip", "100 strips / pack", "pack"])
    book.save(path)
    register_file("file-stale-selected", path)

    source = ComparisonDataSource(
        id="source-stale-selected",
        name="Stale selected source",
        file_id="file-stale-selected",
        sheet_name="Items",
        header_row=1,
        first_data_row=3,
        selected_row_numbers=[4, 5],
        ignored_row_numbers=[],
        row_selection_mode="auto_detected",
        fields=[],
    )

    preview = preview_data_source(source)

    assert preview.data_source.selected_row_numbers == [3, 4, 5]
