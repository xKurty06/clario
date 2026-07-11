from functools import lru_cache
from pathlib import Path
from typing import Any

import pandas as pd

from app.extractors.csv_extractor import CsvExtractor
from app.extractors.excel_extractor import ExcelExtractor
from app.models.file_models import SheetInfo
from app.normalizers.text_normalizer import normalize_text

HEADER_TERMS = ("item", "description", "particular", "specification", "qty", "quantity", "unit", "cost", "price", "amount", "total")
BLANK_HEADER_LABEL = "Blank header"
SHEET_PREVIEW_ROW_LIMIT = 60


def _clean_header_value(value: object) -> str:
    if pd.isna(value):
        return ""
    text = normalize_text(value)
    return "" if text.casefold() == "nan" else text


def clean_headers(values: list[object]) -> list[str]:
    headers: list[str] = []
    counts: dict[str, int] = {}
    for value in values:
        base = _clean_header_value(value) or BLANK_HEADER_LABEL
        count = counts.get(base.casefold(), 0) + 1
        counts[base.casefold()] = count
        headers.append(base if count == 1 else f"{base} ({count})")
    return headers


@lru_cache(maxsize=128)
def _merged_ranges(path: Path, sheet_name: str) -> tuple[tuple[int, int, int, int], ...]:
    """Return zero-based, end-exclusive merged ranges."""
    if path.suffix.lower() == ".xls":
        import xlrd
        workbook = xlrd.open_workbook(path, formatting_info=True)
        return tuple(workbook.sheet_by_name(sheet_name).merged_cells)
    if path.suffix.lower() == ".xlsx":
        import openpyxl
        workbook = openpyxl.load_workbook(path, read_only=False, data_only=True)
        try:
            worksheet = workbook[sheet_name]
            return tuple((cell.min_row - 1, cell.max_row, cell.min_col - 1, cell.max_col) for cell in worksheet.merged_cells.ranges)
        finally:
            workbook.close()
    return ()


def _pad_values(values: list[Any], column_count: int) -> list[Any]:
    if len(values) >= column_count:
        return values[:column_count]
    return values + [None] * (column_count - len(values))


def _selected_header_values(path: Path, sheet_name: str, frame: pd.DataFrame, header_row: int) -> list[str]:
    """Read the selected header row exactly as Excel shows it.

    Pandas and workbook engines can sometimes make merged headers look duplicated.
    For header detection we need the real row values, especially for non-anchor
    merged cells that look blank in Excel.
    """
    row_index = header_row - 1
    column_count = len(frame.columns)
    if row_index < 0 or row_index >= len(frame) or column_count <= 0:
        return []

    suffix = path.suffix.lower()
    values: list[Any]
    try:
        if suffix == ".xlsx":
            import openpyxl
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
            try:
                worksheet = workbook[sheet_name]
                values = [cell.value for cell in next(worksheet.iter_rows(min_row=header_row, max_row=header_row, max_col=column_count))]
            finally:
                workbook.close()
            return [_clean_header_value(value) for value in _pad_values(values, column_count)]
        if suffix == ".xls":
            import xlrd
            workbook = xlrd.open_workbook(path, on_demand=True)
            try:
                worksheet = workbook.sheet_by_name(sheet_name)
                values = worksheet.row_values(row_index, 0, min(column_count, worksheet.ncols)) if row_index < worksheet.nrows else []
            finally:
                workbook.release_resources()
            return [_clean_header_value(value) for value in _pad_values(values, column_count)]
    except Exception:
        pass

    return [_clean_header_value(value) for value in frame.iloc[row_index].tolist()]


def composite_headers(path: Path, sheet_name: str, frame: pd.DataFrame, header_row: int) -> list[str]:
    """Combine split headings ending at the user-selected header row.

    Blank cells on the selected header row intentionally stay blank. This avoids
    showing duplicated merged labels for columns that look empty in Excel.
    Parent merged headings are only used when the selected header row has its own
    visible child header for that column.
    """
    end_index = header_row - 1
    if end_index < 0 or end_index >= len(frame):
        raise ValueError("Header row is outside the worksheet")
    selected_header_values = _selected_header_values(path, sheet_name, frame, header_row)
    ranges = _merged_ranges(path, sheet_name)
    starts = [row_start for row_start, row_end, _, _ in ranges if row_start <= end_index < row_end]
    start_index = min(starts) if starts else end_index
    # Limit accidental inclusion of document titles while allowing common 2-5 row headers.
    start_index = max(start_index, end_index - 5)
    matrix = frame.iloc[start_index:end_index + 1].copy()
    for row_start, row_end, col_start, col_end in ranges:
        if row_end <= start_index or row_start > end_index:
            continue
        anchor = frame.iat[row_start, col_start]
        for row in range(max(row_start, start_index), min(row_end, end_index + 1)):
            for column in range(col_start, min(col_end, len(frame.columns))):
                matrix.iat[row - start_index, column] = anchor
    combined: list[str] = []
    for column in range(len(frame.columns)):
        selected_header_text = selected_header_values[column] if column < len(selected_header_values) else ""
        if not selected_header_text:
            combined.append("")
            continue

        parts: list[str] = []
        for value in matrix.iloc[:, column].tolist():
            text = _clean_header_value(value)
            if text and text.casefold() not in {part.casefold() for part in parts}:
                parts.append(text)
        combined.append(" / ".join(parts) if parts else selected_header_text)
    return clean_headers(combined)


def detect_header_index(frame: pd.DataFrame) -> int:
    if frame.empty:
        return 0
    best_index, best_score = 0, -1
    for index in range(min(len(frame), SHEET_PREVIEW_ROW_LIMIT)):
        values = [normalize_text(value, True) for value in frame.iloc[index].tolist() if not pd.isna(value)]
        non_empty = len(values)
        keyword_hits = sum(any(term in value for term in HEADER_TERMS) for value in values)
        distinct_hits = len({term for term in HEADER_TERMS if any(term in value for value in values)})
        score = keyword_hits * 4 + distinct_hits * 2 + min(non_empty, 8)
        if non_empty >= 2 and score > best_score:
            best_index, best_score = index, score
    return best_index


def inspect_header(path: Path, sheet_name: str, header_row: int) -> tuple[list[str], list[dict[str, object]]]:
    frames = CsvExtractor().extract(path) if path.suffix.lower() == ".csv" else ExcelExtractor().extract(path)
    if sheet_name not in frames:
        raise ValueError("Sheet not found")
    frame = frames[sheet_name]
    index = header_row - 1
    if index < 0 or index >= len(frame):
        raise ValueError("Header row is outside the worksheet")
    headers = composite_headers(path, sheet_name, frame, header_row)
    samples = [{headers[i]: (None if pd.isna(value) else str(value)) for i, value in enumerate(row.tolist())} for _, row in frame.iloc[index + 1:index + 6].iterrows()]
    return headers, samples


def _frame_from_preview_rows(rows: list[list[Any]] | list[tuple[Any, ...]], column_count: int) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=range(max(column_count, 0)))
    frame = pd.DataFrame([_pad_values(list(row), column_count) for row in rows])
    if column_count and len(frame.columns) < column_count:
        frame = frame.reindex(columns=range(column_count))
    return frame


def _make_basic_sheet_info(name: str, row_count: int, column_count: int, rows: list[list[Any]] | list[tuple[Any, ...]]) -> SheetInfo:
    frame = _frame_from_preview_rows(rows, column_count)
    has_real_values = not frame.empty and any(any(not pd.isna(value) and normalize_text(value) for value in row) for row in frame.values.tolist())
    detected_header_row = detect_header_index(frame) + 1 if has_real_values else 1
    return SheetInfo(
        name=name,
        row_count=max(row_count, 0),
        column_count=max(column_count, 0),
        detected_header_row=detected_header_row,
        headers=[],
        sample_rows=[],
    )


def _inspect_xlsx_sheets(path: Path) -> list[SheetInfo]:
    import openpyxl
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        result: list[SheetInfo] = []
        for worksheet in workbook.worksheets:
            row_count = worksheet.max_row or 0
            column_count = worksheet.max_column or 0
            rows = list(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=min(row_count, SHEET_PREVIEW_ROW_LIMIT),
                    max_col=column_count,
                    values_only=True,
                )
            ) if row_count and column_count else []
            result.append(_make_basic_sheet_info(worksheet.title, row_count, column_count, rows))
        return result
    finally:
        workbook.close()


def _inspect_xls_sheets(path: Path) -> list[SheetInfo]:
    import xlrd
    workbook = xlrd.open_workbook(path, on_demand=True)
    try:
        result: list[SheetInfo] = []
        for name in workbook.sheet_names():
            worksheet = workbook.sheet_by_name(name)
            rows = [worksheet.row_values(row_index, 0, worksheet.ncols) for row_index in range(min(worksheet.nrows, SHEET_PREVIEW_ROW_LIMIT))]
            result.append(_make_basic_sheet_info(name, worksheet.nrows, worksheet.ncols, rows))
        return result
    finally:
        workbook.release_resources()


def _inspect_csv_sheet(path: Path) -> list[SheetInfo]:
    try:
        frame = pd.read_csv(path, header=None, dtype=object, encoding="utf-8-sig", nrows=SHEET_PREVIEW_ROW_LIMIT)
    except UnicodeDecodeError:
        frame = pd.read_csv(path, header=None, dtype=object, encoding="latin-1", nrows=SHEET_PREVIEW_ROW_LIMIT)
    try:
        with path.open("r", encoding="utf-8-sig", errors="ignore") as handle:
            row_count = sum(1 for _ in handle)
    except OSError:
        row_count = len(frame)
    return [_make_basic_sheet_info("CSV", row_count, len(frame.columns), frame.values.tolist())]


def inspect_sheets(path: Path) -> list[SheetInfo]:
    """Return fast sheet metadata for upload.

    Upload should feel quick, so this intentionally avoids full pandas extraction,
    merged-header expansion, and per-sheet samples. Deep header/sample inspection
    still happens later for the chosen sheet through inspect_header/preview.
    """
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _inspect_csv_sheet(path)
    if suffix == ".xlsx":
        return _inspect_xlsx_sheets(path)
    if suffix == ".xls":
        return _inspect_xls_sheets(path)
    return []
